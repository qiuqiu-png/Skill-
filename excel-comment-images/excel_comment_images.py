#!/usr/bin/env python3
"""
Excel 批注插图工具
将图片以批注（备注）形式插入到 Excel 单元格中
鼠标悬停在单元格上即可预览图片
"""

import os
import argparse
import zipfile
import shutil
from openpyxl import load_workbook
from openpyxl.comments import Comment


def find_matching_image(target_name, image_files):
    """在图片列表中查找文件名匹配的图片（不含扩展名）"""
    for f in image_files:
        if os.path.splitext(f)[0] == target_name:
            return f
    return None


def col_to_index(col_letter):
    """列字母转为 0-based 索引：A→0, B→1, C→2"""
    result = 0
    for c in col_letter.upper():
        result = result * 26 + (ord(c) - ord('A'))
    return result


def generate_vml(shapes_info, width_pt, height_pt):
    """生成标准 VML，使用 v:/o:/x:/r: 命名空间前缀"""
    parts = []
    parts.append(
        '<xml xmlns:v="urn:schemas-microsoft-com:vml"'
        ' xmlns:o="urn:schemas-microsoft-com:office:office"'
        ' xmlns:x="urn:schemas-microsoft-com:office:excel"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">')
    parts.append(
        '<o:shapelayout v:ext="edit">'
        '<o:idmap v:ext="edit" data="1"/>'
        '</o:shapelayout>')
    parts.append(
        '<v:shapetype id="_x0000_t202" coordsize="21600,21600" o:spt="202"'
        ' path="m,l,21600r21600,l21600,xe">'
        '<v:stroke joinstyle="miter"/>'
        '<v:path gradientshapeok="t" o:connecttype="rect"/>'
        '</v:shapetype>')

    for i, (row, col, rid) in enumerate(shapes_info):
        parts.append(
            f'<v:shape id="_x0000_s{1025 + i}" type="#_x0000_t202"'
            f' style="position:absolute;margin-left:59.25pt;margin-top:1.5pt;'
            f'width:{width_pt}pt;height:{height_pt}pt;'
            f'z-index:{i + 1};visibility:hidden"'
            f' o:insetmode="auto">'
            f'<v:fill type="frame" o:detectmouseclick="t" recolor="t" r:id="{rid}"/>'
            f'<v:shadow on="t" color="black" obscured="t"/>'
            f'<v:path o:connecttype="none"/>'
            f'<v:textbox style="mso-direction-alt:auto">'
            f'<div style="text-align:left"></div>'
            f'</v:textbox>'
            f'<x:ClientData ObjectType="Note">'
            f'<x:MoveWithCells/>'
            f'<x:SizeWithCells/>'
            f'<x:AutoFill>False</x:AutoFill>'
            f'<x:Row>{row}</x:Row>'
            f'<x:Column>{col}</x:Column>'
            f'</x:ClientData>'
            f'</v:shape>')

    parts.append('</xml>')
    return '\n'.join(parts)


def make_rels_xml(images):
    """生成 VML 图片关系文件"""
    lines = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
    ]
    for rid, media_path, _ in images:
        name = os.path.basename(media_path)
        lines.append(
            f'  <Relationship Id="{rid}" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
            f'Target="../media/{name}"/>')
    lines.append('</Relationships>')
    return '\n'.join(lines)


def update_content_types(ct_xml, images):
    """确保 [Content_Types].xml 包含所需的图片类型声明"""
    ext_map = {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
        '.png': 'image/png', '.gif': 'image/gif', '.bmp': 'image/bmp'
    }
    for _, media_path, _ in images:
        ext = os.path.splitext(media_path)[1].lower()
        content_type = ext_map.get(ext, 'image/png')
        ext_no_dot = ext[1:]
        if f'Extension="{ext_no_dot}"' not in ct_xml:
            ct_xml = ct_xml.replace(
                '</Types>',
                f'  <Default Extension="{ext_no_dot}" ContentType="{content_type}"/>\n</Types>')
    return ct_xml


def inject_images_into_xlsx(xlsx_path, image_map, width_pt, height_pt):
    """后处理 xlsx：用标准 VML 替换 openpyxl 生成的非标准 VML，注入图片"""
    temp_path = xlsx_path + ".tmp"

    # 准备形状和图片列表
    shapes_info = []
    all_images = []
    counter = 1
    for (row, col), src_path in sorted(image_map.items()):
        rid = f"rId{counter}"
        ext = os.path.splitext(src_path)[1].lower()
        media_name = f"comment_img{counter}{ext}"
        shapes_info.append((row, col, rid))
        all_images.append((rid, f"xl/media/{media_name}", src_path))
        counter += 1

    # 生成标准 VML
    new_vml = generate_vml(shapes_info, width_pt, height_pt)

    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        file_data = {}

        for item in zin.infolist():
            fn = item.filename
            data = zin.read(fn)

            if fn.endswith('.vml') and '/drawings/' in fn:
                file_data[fn] = new_vml.encode('utf-8')
                d = os.path.dirname(fn)
                b = os.path.basename(fn)
                rels_path = f"{d}/_rels/{b}.rels"
                file_data[rels_path] = make_rels_xml(all_images).encode('utf-8')
            else:
                file_data[fn] = data

        # 更新 Content_Types
        ct_key = '[Content_Types].xml'
        if all_images and ct_key in file_data:
            ct_text = file_data[ct_key].decode('utf-8')
            ct_text = update_content_types(ct_text, all_images)
            file_data[ct_key] = ct_text.encode('utf-8')

        # 写入新 xlsx
        with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for fn, data in file_data.items():
                zout.writestr(fn, data)
            for _, media_path, src_path in all_images:
                with open(src_path, 'rb') as f:
                    zout.writestr(media_path, f.read())

    shutil.move(temp_path, xlsx_path)


def insert_comment_images(excel_path, images_folder, output_path=None,
                          name_col='C', start_row=1, end_row=65536,
                          width=240, height=160):
    """
    Excel 批注插图主函数

    参数:
        excel_path: Excel 文件路径
        images_folder: 图片文件夹路径
        output_path: 输出路径（None 则自动生成）
        name_col: 文件名所在列字母
        start_row/end_row: 行范围
        width/height: 批注大小（pt）
    """
    if output_path is None:
        base, ext = os.path.splitext(excel_path)
        output_path = f"{base}_批注{ext}"

    print(f"正在加载 Excel: {excel_path}")
    try:
        wb = load_workbook(excel_path)
    except Exception as e:
        print(f"无法加载 Excel 文件: {e}")
        return False

    ws = wb.active

    image_files = [f for f in os.listdir(images_folder)
                   if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif'))]

    if not image_files:
        print(f"图片文件夹中未找到图片: {images_folder}")
        return False

    print(f"找到 {len(image_files)} 个图片文件")
    print(f"模式: 批注插图 | 文件名列: {name_col} | 行范围: {start_row}-{end_row}")
    print(f"批注大小: {width} x {height} pt")
    print("-" * 60)

    max_row = min(end_row, ws.max_row)
    col_idx = col_to_index(name_col)

    image_map = {}
    success_count = 0
    not_found = []

    for row in range(start_row, max_row + 1):
        cell = ws[f"{name_col}{row}"]
        if not cell.value:
            continue

        name = str(cell.value).strip()
        if '.' in name:
            name = os.path.splitext(name)[0]

        matched = find_matching_image(name, image_files)
        if matched:
            cell.comment = Comment(" ", "")
            image_map[(row - 1, col_idx)] = os.path.join(images_folder, matched)
            print(f"  {matched} -> {name_col}{row}")
            success_count += 1
        else:
            not_found.append(name)

    if success_count == 0:
        print("没有匹配到任何图片")
        return False

    print(f"\n正在保存并注入图片到批注...")
    wb.save(output_path)
    inject_images_into_xlsx(output_path, image_map, width, height)

    print("-" * 60)
    print(f"处理完成！成功插入: {success_count} 张批注图片")
    if not_found:
        s = ', '.join(not_found[:10])
        if len(not_found) > 10:
            s += f" ...等共 {len(not_found)} 项"
        print(f"未匹配: {s}")
    print(f"输出文件: {output_path}")
    return True


def main():
    parser = argparse.ArgumentParser(
        description='Excel 批注插图工具：将图片以批注形式插入单元格',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  %(prog)s --excel products.xlsx --images images/ --name-col C --start-row 2
  %(prog)s --excel list.xlsx --images photos/ --output result.xlsx --width 300 --height 200
        """
    )

    parser.add_argument('--excel', required=True, help='Excel 文件路径')
    parser.add_argument('--images', required=True, help='图片文件夹路径')
    parser.add_argument('--output', help='输出文件路径（默认：原文件名_批注.xlsx）')
    parser.add_argument('--name-col', default='C', help='文件名所在列（默认：C）')
    parser.add_argument('--start-row', type=int, default=1, help='起始行号（默认：1）')
    parser.add_argument('--end-row', type=int, default=65536, help='结束行号（默认：65536）')
    parser.add_argument('--width', type=int, default=240, help='批注宽度 pt（默认：240）')
    parser.add_argument('--height', type=int, default=160, help='批注高度 pt（默认：160）')

    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"错误：Excel 文件不存在: {args.excel}")
        return 1
    if not os.path.exists(args.images):
        print(f"错误：图片文件夹不存在: {args.images}")
        return 1

    success = insert_comment_images(
        excel_path=args.excel,
        images_folder=args.images,
        output_path=args.output,
        name_col=args.name_col,
        start_row=args.start_row,
        end_row=args.end_row,
        width=args.width,
        height=args.height,
    )
    return 0 if success else 1


if __name__ == "__main__":
    exit(main())
