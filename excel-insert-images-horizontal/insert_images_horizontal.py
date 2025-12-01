#!/usr/bin/env python3
"""
Excel 横向插图工具
将图片按行插入 Excel 表格
"""

import os
import argparse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage


def insert_images_horizontal(excel_path, images_folder, output_excel_path=None,
                             start_row=1, end_row=30, name_col='B'):
    """
    将图片按行插入到 Excel 表格中

    参数:
        excel_path: 原始 Excel 文件路径
        images_folder: 包含图片的文件夹路径
        output_excel_path: 输出 Excel 文件路径（如果为 None，则覆盖原文件）
        start_row: 起始行号（默认第1行）
        end_row: 结束行号（默认第30行）
        name_col: 文件名列（默认'B'）
    """
    # 如果未指定输出路径，则覆盖原文件
    if output_excel_path is None:
        output_excel_path = excel_path

    # 加载 Excel 文件
    print(f"正在加载 Excel: {excel_path}")
    wb = load_workbook(excel_path)

    # 选择第一个工作表
    if wb.sheetnames:
        sheet = wb.active
    else:
        print("✗ Excel 文件中没有工作表")
        return False

    # 获取所有图片文件
    image_files = [f for f in os.listdir(images_folder)
                   if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif'))]

    if not image_files:
        print(f"✗ 在图片文件夹中未找到任何图片: {images_folder}")
        return False

    print(f"找到 {len(image_files)} 个图片文件")
    print(f"处理范围: {start_row} 到 {end_row} 行")
    print("-" * 60)

    success_count = 0
    skip_count = 0

    # 遍历指定行范围
    for row in range(start_row, end_row + 1):
        # 获取文件名列的值
        cell = sheet[f"{name_col}{row}"]
        if not cell.value:
            print(f"⚠ 第 {row} 行 {name_col} 列为空，跳过")
            skip_count += 1
            continue

        # 获取文件名（不带扩展名）
        target_filename = str(cell.value)
        if '.' in target_filename:
            target_filename = os.path.splitext(target_filename)[0]

        # 查找匹配的图片文件
        matched_image = None
        for img_file in image_files:
            img_name = os.path.splitext(img_file)[0]
            if img_name == target_filename:
                matched_image = img_file
                break

        if not matched_image:
            print(f"⚠ 第 {row} 行未找到匹配图片: {target_filename}")
            skip_count += 1
            continue

        # 构建图片完整路径
        image_path = os.path.join(images_folder, matched_image)

        try:
            # 创建 Excel 图片对象
            excel_img = ExcelImage(image_path)

            # 计算图片插入位置（插入到 name_col 的下一列）
            col_letter = chr(ord(name_col) + 1)
            cell_ref = f"{col_letter}{row}"

            # 添加图片到工作表
            sheet.add_image(excel_img, cell_ref)
            print(f"✓ 第 {row} 行: {matched_image} → {cell_ref}")
            success_count += 1

        except Exception as e:
            print(f"✗ 第 {row} 行插入失败: {e}")
            skip_count += 1

    # 保存 Excel 文件
    print("-" * 60)
    print("正在保存文件...")
    wb.save(output_excel_path)

    print(f"✓ 处理完成！")
    print(f"  成功插入: {success_count} 张图片")
    print(f"  跳过: {skip_count} 行")
    print(f"  输出文件: {output_excel_path}")

    return True


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='Excel 横向插图工具：将图片按行插入 Excel 表格',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 基本用法
  %(prog)s --excel products.xlsx --images images/

  # 完整参数
  %(prog)s --excel data.xlsx --images photos/ --output result.xlsx --start-row 1 --end-row 100 --name-col A
        """
    )

    parser.add_argument('--excel', required=True, help='Excel 文件路径')
    parser.add_argument('--images', required=True, help='图片文件夹路径')
    parser.add_argument('--output', help='输出文件路径（默认覆盖原文件）')
    parser.add_argument('--start-row', type=int, default=1, help='起始行号（默认：1）')
    parser.add_argument('--end-row', type=int, default=30, help='结束行号（默认：30）')
    parser.add_argument('--name-col', default='B', help='文件名所在列（默认：B）')

    args = parser.parse_args()

    # 验证参数
    if not os.path.exists(args.excel):
        print(f"✗ 错误：Excel 文件不存在: {args.excel}")
        return 1

    if not os.path.exists(args.images):
        print(f"✗ 错误：图片文件夹不存在: {args.images}")
        return 1

    if args.start_row < 1:
        print(f"✗ 错误：起始行号必须 >= 1")
        return 1

    if args.end_row < args.start_row:
        print(f"✗ 错误：结束行号必须 >= 起始行号")
        return 1

    print("=" * 60)
    print("Excel 横向插图工具")
    print("=" * 60)
    print()

    # 执行插入
    success = insert_images_horizontal(
        excel_path=args.excel,
        images_folder=args.images,
        output_excel_path=args.output,
        start_row=args.start_row,
        end_row=args.end_row,
        name_col=args.name_col
    )

    return 0 if success else 1


if __name__ == "__main__":
    exit(main())
