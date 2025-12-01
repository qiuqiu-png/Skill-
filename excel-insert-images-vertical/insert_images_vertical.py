#!/usr/bin/env python3
"""
Excel 竖向插图工具
将图片按列插入 Excel 表格
"""

import os
import argparse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils import get_column_letter


def insert_images_vertical(excel_path, images_folder, output_excel_path=None,
                           name_row=1, img_row=2, start_col=1, end_col=26, img_height=300):
    """
    将图片按列插入到 Excel 表格中

    参数:
        excel_path: 原始 Excel 文件路径
        images_folder: 包含图片的文件夹路径
        output_excel_path: 输出 Excel 文件路径（如果为 None，则覆盖原文件）
        name_row: 文件名所在行（默认1）
        img_row: 图片插入行（默认2）
        start_col: 起始列号（默认1，对应A列）
        end_col: 结束列号（默认26，对应Z列）
        img_height: 图片高度（像素，默认300）
    """
    # 如果未指定输出路径，则覆盖原文件
    if output_excel_path is None:
        output_excel_path = excel_path

    # 加载 Excel 文件
    print(f"正在加载 Excel: {excel_path}")
    try:
        wb = load_workbook(excel_path)
    except Exception as e:
        print(f"✗ 无法加载 Excel 文件: {e}")
        return False

    # 选择第一个工作表
    if wb.sheetnames:
        sheet = wb.active
    else:
        print("✗ Excel 文件中没有工作表")
        return False

    # 获取所有图片文件
    image_files = [f for f in os.listdir(images_folder)
                   if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif'))]

    if not image_files:
        print(f"✗ 在图片文件夹中未找到任何图片: {images_folder}")
        return False

    print(f"找到 {len(image_files)} 个图片文件")
    print(f"处理范围: 第 {start_col} 到 {end_col} 列")
    print(f"图片高度: {img_height} 像素")
    print("-" * 60)

    # 调整图片行的高度
    sheet.row_dimensions[img_row].height = img_height * 0.75

    success_count = 0
    skip_count = 0

    # 遍历指定列范围
    for col in range(start_col, end_col + 1):
        col_letter = get_column_letter(col)

        # 获取文件名列的单元格值
        cell = sheet[f"{col_letter}{name_row}"]
        if not cell.value:
            skip_count += 1
            continue

        # 获取文件名（不带扩展名）
        target_filename = str(cell.value)
        if '.' in target_filename:
            target_filename = os.path.splitext(target_filename)[0]

        # 查找匹配的图片文件
        matched_image = None
        for img_file in image_files:
            img_name = os.path.splitext(img_file)[0]
            if img_name == target_filename:
                matched_image = img_file
                break

        if not matched_image:
            print(f"⚠ {col_letter}{name_row} 未找到匹配图片: {target_filename}")
            skip_count += 1
            continue

        # 构建图片完整路径
        image_path = os.path.join(images_folder, matched_image)

        try:
            # 创建 Excel 图片对象
            excel_img = ExcelImage(image_path)

            # 计算图片尺寸（保持比例，按指定高度缩放）
            original_width, original_height = excel_img.width, excel_img.height
            scale = img_height / original_height
            new_width = original_width * scale
            new_height = img_height

            # 设置图片尺寸
            excel_img.width = new_width
            excel_img.height = new_height

            # 调整列宽以适应图片宽度
            required_width = new_width / 7
            if sheet.column_dimensions[col_letter].width < required_width:
                sheet.column_dimensions[col_letter].width = required_width

            # 图片插入位置
            cell_ref = f"{col_letter}{img_row}"

            # 设置图片偏移，使其水平居中
            excel_img.anchor = cell_ref
            current_col_width = sheet.column_dimensions[col_letter].width
            offset_x = (current_col_width * 7 - new_width) / 2
            if offset_x > 0:
                excel_img.dx = pixels_to_EMU(offset_x)
            excel_img.dy = 0

            # 添加图片到工作表
            sheet.add_image(excel_img, cell_ref)
            print(f"✓ {col_letter} 列: {matched_image} → {cell_ref}")
            success_count += 1

        except Exception as e:
            print(f"✗ {col_letter} 列插入失败: {e}")
            skip_count += 1

    # 保存 Excel 文件
    print("-" * 60)
    print("正在保存文件...")
    try:
        wb.save(output_excel_path)
    except Exception as e:
        print(f"✗ 保存文件失败: {e}")
        return False

    print(f"✓ 处理完成！")
    print(f"  成功插入: {success_count} 张图片")
    print(f"  跳过: {skip_count} 列")
    print(f"  输出文件: {output_excel_path}")

    return True


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='Excel 竖向插图工具：将图片按列插入 Excel 表格',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 基本用法
  %(prog)s --excel comparison.xlsx --images images/

  # 完整参数
  %(prog)s --excel data.xlsx --images photos/ --output result.xlsx --name-row 1 --img-row 2 --start-col 1 --end-col 20 --img-height 400
        """
    )

    parser.add_argument('--excel', required=True, help='Excel 文件路径')
    parser.add_argument('--images', required=True, help='图片文件夹路径')
    parser.add_argument('--output', help='输出文件路径（默认覆盖原文件）')
    parser.add_argument('--name-row', type=int, default=1, help='文件名所在行（默认：1）')
    parser.add_argument('--img-row', type=int, default=2, help='图片插入行（默认：2）')
    parser.add_argument('--start-col', type=int, default=1, help='起始列号，1=A列（默认：1）')
    parser.add_argument('--end-col', type=int, default=26, help='结束列号，26=Z列（默认：26）')
    parser.add_argument('--img-height', type=int, default=300, help='图片高度（像素，默认：300）')

    args = parser.parse_args()

    # 验证参数
    if not os.path.exists(args.excel):
        print(f"✗ 错误：Excel 文件不存在: {args.excel}")
        return 1

    if not os.path.exists(args.images):
        print(f"✗ 错误：图片文件夹不存在: {args.images}")
        return 1

    if args.name_row < 1 or args.img_row < 1:
        print(f"✗ 错误：行号必须 >= 1")
        return 1

    if args.start_col < 1 or args.end_col < 1:
        print(f"✗ 错误：列号必须 >= 1")
        return 1

    if args.end_col < args.start_col:
        print(f"✗ 错误：结束列号必须 >= 起始列号")
        return 1

    if args.img_height < 50:
        print(f"✗ 错误：图片高度太小，建议 >= 50 像素")
        return 1

    print("=" * 60)
    print("Excel 竖向插图工具")
    print("=" * 60)
    print()

    # 执行插入
    success = insert_images_vertical(
        excel_path=args.excel,
        images_folder=args.images,
        output_excel_path=args.output,
        name_row=args.name_row,
        img_row=args.img_row,
        start_col=args.start_col,
        end_col=args.end_col,
        img_height=args.img_height
    )

    return 0 if success else 1


if __name__ == "__main__":
    exit(main())
