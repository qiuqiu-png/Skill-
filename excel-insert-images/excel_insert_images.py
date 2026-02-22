#!/usr/bin/env python3
"""
Excel 批量插图工具
支持横向（按行）和竖向（按列）两种模式
"""

import os
import argparse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils import get_column_letter


def find_matching_image(target_filename, image_files):
    """在图片列表中查找匹配的文件"""
    for img_file in image_files:
        img_name = os.path.splitext(img_file)[0]
        if img_name == target_filename:
            return img_file
    return None


def get_cell_filename(cell):
    """从单元格获取文件名（去除扩展名）"""
    if not cell.value:
        return None
    target = str(cell.value)
    if '.' in target:
        target = os.path.splitext(target)[0]
    return target


def insert_horizontal(sheet, images_folder, image_files, start_row, end_row, name_col):
    """横向模式：按行插入图片"""
    print(f"模式: 横向（按行）")
    print(f"处理范围: 第 {start_row} 到 {end_row} 行, 文件名列: {name_col}")
    print("-" * 60)

    success_count = 0
    skip_count = 0

    for row in range(start_row, end_row + 1):
        cell = sheet[f"{name_col}{row}"]
        target_filename = get_cell_filename(cell)

        if not target_filename:
            skip_count += 1
            continue

        matched_image = find_matching_image(target_filename, image_files)
        if not matched_image:
            print(f"  {name_col}{row} 未找到匹配图片: {target_filename}")
            skip_count += 1
            continue

        image_path = os.path.join(images_folder, matched_image)
        try:
            excel_img = ExcelImage(image_path)
            col_letter = chr(ord(name_col) + 1)
            cell_ref = f"{col_letter}{row}"
            sheet.add_image(excel_img, cell_ref)
            print(f"  {matched_image} -> {cell_ref}")
            success_count += 1
        except Exception as e:
            print(f"  第 {row} 行插入失败: {e}")
            skip_count += 1

    return success_count, skip_count


def insert_vertical(sheet, images_folder, image_files, name_row, img_row,
                    start_col, end_col, img_height):
    """竖向模式：按列插入图片"""
    print(f"模式: 竖向（按列）")
    print(f"处理范围: 第 {start_col} 到 {end_col} 列, 图片高度: {img_height}px")
    print("-" * 60)

    sheet.row_dimensions[img_row].height = img_height * 0.75

    success_count = 0
    skip_count = 0

    for col in range(start_col, end_col + 1):
        col_letter = get_column_letter(col)
        cell = sheet[f"{col_letter}{name_row}"]
        target_filename = get_cell_filename(cell)

        if not target_filename:
            skip_count += 1
            continue

        matched_image = find_matching_image(target_filename, image_files)
        if not matched_image:
            print(f"  {col_letter}{name_row} 未找到匹配图片: {target_filename}")
            skip_count += 1
            continue

        image_path = os.path.join(images_folder, matched_image)
        try:
            excel_img = ExcelImage(image_path)

            # 按指定高度等比缩放
            scale = img_height / excel_img.height
            new_width = excel_img.width * scale
            excel_img.width = new_width
            excel_img.height = img_height

            # 调整列宽
            required_width = new_width / 7
            if sheet.column_dimensions[col_letter].width < required_width:
                sheet.column_dimensions[col_letter].width = required_width

            cell_ref = f"{col_letter}{img_row}"
            excel_img.anchor = cell_ref

            # 水平居中
            current_col_width = sheet.column_dimensions[col_letter].width
            offset_x = (current_col_width * 7 - new_width) / 2
            if offset_x > 0:
                excel_img.dx = pixels_to_EMU(offset_x)
            excel_img.dy = 0

            sheet.add_image(excel_img, cell_ref)
            print(f"  {matched_image} -> {cell_ref}")
            success_count += 1
        except Exception as e:
            print(f"  {col_letter} 列插入失败: {e}")
            skip_count += 1

    return success_count, skip_count


def insert_images(excel_path, images_folder, mode='horizontal', output_path=None,
                  start_row=1, end_row=30, name_col='B',
                  name_row=1, img_row=2, start_col=1, end_col=26, img_height=300):
    """
    Excel 批量插图主函数

    参数:
        excel_path: Excel 文件路径
        images_folder: 图片文件夹路径
        mode: 插入模式 ('horizontal' 或 'vertical')
        output_path: 输出路径（None 则覆盖原文件）
        start_row/end_row/name_col: 横向模式参数
        name_row/img_row/start_col/end_col/img_height: 竖向模式参数
    """
    if output_path is None:
        base, ext = os.path.splitext(excel_path)
        output_path = f"{base}_output{ext}"

    print(f"正在加载 Excel: {excel_path}")
    try:
        wb = load_workbook(excel_path)
    except Exception as e:
        print(f"无法加载 Excel 文件: {e}")
        return False

    if not wb.sheetnames:
        print("Excel 文件中没有工作表")
        return False

    sheet = wb.active

    image_files = [f for f in os.listdir(images_folder)
                   if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif'))]

    if not image_files:
        print(f"在图片文件夹中未找到任何图片: {images_folder}")
        return False

    print(f"找到 {len(image_files)} 个图片文件")

    if mode == 'horizontal':
        success, skip = insert_horizontal(sheet, images_folder, image_files,
                                          start_row, end_row, name_col)
    else:
        success, skip = insert_vertical(sheet, images_folder, image_files,
                                        name_row, img_row, start_col, end_col, img_height)

    print("-" * 60)
    print("正在保存文件...")
    try:
        wb.save(output_path)
    except Exception as e:
        print(f"保存文件失败: {e}")
        return False

    print(f"处理完成！成功插入: {success} 张, 跳过: {skip}")
    print(f"输出文件: {output_path}")
    return True


def main():
    parser = argparse.ArgumentParser(
        description='Excel 批量插图工具：支持横向和竖向两种模式',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 横向插入（按行）
  %(prog)s --excel products.xlsx --images images/ --mode horizontal --name-col B

  # 竖向插入（按列）
  %(prog)s --excel comparison.xlsx --images images/ --mode vertical --img-height 400
        """
    )

    parser.add_argument('--excel', required=True, help='Excel 文件路径')
    parser.add_argument('--images', required=True, help='图片文件夹路径')
    parser.add_argument('--mode', choices=['horizontal', 'vertical'], default='horizontal',
                        help='插入模式（默认：horizontal）')
    parser.add_argument('--output', help='输出文件路径（默认在原文件名后加 _output）')

    # 横向模式参数
    parser.add_argument('--name-col', default='B', help='[横向] 文件名所在列（默认：B）')
    parser.add_argument('--start-row', type=int, default=1, help='[横向] 起始行号（默认：1）')
    parser.add_argument('--end-row', type=int, default=30, help='[横向] 结束行号（默认：30）')

    # 竖向模式参数
    parser.add_argument('--name-row', type=int, default=1, help='[竖向] 文件名所在行（默认：1）')
    parser.add_argument('--img-row', type=int, default=2, help='[竖向] 图片插入行（默认：2）')
    parser.add_argument('--start-col', type=int, default=1, help='[竖向] 起始列号（默认：1）')
    parser.add_argument('--end-col', type=int, default=26, help='[竖向] 结束列号（默认：26）')
    parser.add_argument('--img-height', type=int, default=300, help='[竖向] 图片高度像素（默认：300）')

    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"错误：Excel 文件不存在: {args.excel}")
        return 1
    if not os.path.exists(args.images):
        print(f"错误：图片文件夹不存在: {args.images}")
        return 1

    success = insert_images(
        excel_path=args.excel,
        images_folder=args.images,
        mode=args.mode,
        output_path=args.output,
        start_row=args.start_row,
        end_row=args.end_row,
        name_col=args.name_col,
        name_row=args.name_row,
        img_row=args.img_row,
        start_col=args.start_col,
        end_col=args.end_col,
        img_height=args.img_height,
    )
    return 0 if success else 1


if __name__ == "__main__":
    exit(main())
