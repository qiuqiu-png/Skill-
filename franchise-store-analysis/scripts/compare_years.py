#!/usr/bin/env python3
"""
加盟新老店批发毛利对比分析脚本 - 2024年 vs 2025年
支持按年份过滤黄金外采数据，生成完整的同比对比报表
"""

import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path


def merge_excel_files(file_paths):
    """
    合并多个相似的Excel文件

    参数:
        file_paths: 单个文件路径(str)或文件路径列表(list)

    返回:
        合并后的DataFrame
    """
    # 如果是字符串，转换为列表
    if isinstance(file_paths, str):
        file_paths = [file_paths]

    # 如果只有一个文件，直接读取
    if len(file_paths) == 1:
        print(f"      读取单个文件: {file_paths[0]}")
        return pd.read_excel(file_paths[0], sheet_name=0)

    # 合并多个文件
    print(f"      检测到 {len(file_paths)} 个文件，开始合并...")
    dfs = []
    for i, file_path in enumerate(file_paths, 1):
        df = pd.read_excel(file_path, sheet_name=0)
        print(f"      - 文件{i}: {file_path} ({len(df)}行)")
        dfs.append(df)

    merged_df = pd.concat(dfs, ignore_index=True)
    print(f"      ✅ 合并完成，总计 {len(merged_df)} 行")

    return merged_df


def filter_by_year(df, year, time_col='创建时间'):
    """按年份过滤数据"""
    if time_col in df.columns:
        df['年份'] = pd.to_datetime(df[time_col], errors='coerce').dt.year
        return df[df['年份'] == year].copy()
    return df


def process_data(df, org_type_col, new_stores, closed_stores):
    """处理数据：过滤M结尾门店、剔除闭店、标记新老店"""
    # 只保留M结尾的门店
    df = df[df[org_type_col].str.rstrip().str.endswith('M', na=False)]
    
    # 清理组织名称（仅去除*号，保留M）
    df['组织_cleaned'] = df[org_type_col].str.replace(r'[*]+$', '', regex=True)
    
    # 剔除已闭店
    df = df[~df['组织_cleaned'].isin(closed_stores)]
    
    # 标记新店/老店
    df['是否新店'] = df['组织_cleaned'].isin(new_stores)
    
    return df


def analyze_year(year, sales_file, return_file, org_mapping, 
                warehouse_file=None, other_settlement_file=None):
    """分析单个年份的数据"""
    
    print(f"\n{'='*100}")
    print(f"📊 分析{year}年数据")
    print(f"{'='*100}")
    
    # 读取销售数据（支持多文件合并）
    print(f"   读取销售数据...")
    df_sales = merge_excel_files(sales_file)
    df_sales = process_data(df_sales, '组织', org_mapping['new_stores'], org_mapping['closed_stores'])
    df_sales['销售金额'] = pd.to_numeric(df_sales['销售金额'], errors='coerce').fillna(0)
    df_sales['暂估成本'] = pd.to_numeric(df_sales['暂估成本'], errors='coerce').fillna(0)
    print(f"   销售数据: {len(df_sales)}行")
    
    # 读取退货数据（支持多文件合并）
    print(f"   读取退货数据...")
    df_return = merge_excel_files(return_file)
    df_return = process_data(df_return, '组织', org_mapping['new_stores'], org_mapping['closed_stores'])
    df_return['退回金额'] = pd.to_numeric(df_return['退回金额'], errors='coerce').fillna(0)
    df_return['暂估成本'] = pd.to_numeric(df_return['暂估成本'], errors='coerce').fillna(0)
    print(f"   退货数据: {len(df_return)}行")
    
    # 读取入库单（支持多文件合并，按年份过滤）
    df_warehouse = None
    if warehouse_file:
        print(f"   读取入库单数据...")
        df_warehouse = merge_excel_files(warehouse_file)
        df_warehouse = filter_by_year(df_warehouse, year, '创建时间')
        df_warehouse = process_data(df_warehouse, '组织', org_mapping['new_stores'], org_mapping['closed_stores'])
        df_warehouse['总成本'] = pd.to_numeric(df_warehouse['总成本'], errors='coerce').fillna(0)
        print(f"   入库单数据({year}年): {len(df_warehouse)}行")
    
    # 读取其他结算单（支持多文件合并，按年份过滤）
    df_other = None
    if other_settlement_file:
        print(f"   读取其他结算单数据...")
        df_other = merge_excel_files(other_settlement_file)
        df_other = filter_by_year(df_other, year, '创建时间')
        df_other = df_other[
            (df_other['状态'] == '已扣款') &
            (df_other['结算类型'] == '服务费(挂标签）')
        ]
        if '加盟门店' in df_other.columns:
            df_other = process_data(df_other, '加盟门店', org_mapping['new_stores'], org_mapping['closed_stores'])
            df_other['发生金额'] = pd.to_numeric(df_other['发生金额'], errors='coerce').fillna(0)
        print(f"   其他结算单数据({year}年，已扣款): {len(df_other)}行")
    
    # 按品类统计
    categories = ['黄金', '钻石', '爱尚炫', '18K', '翡翠', '其他']
    results = {}
    
    for category in categories:
        sales_cat = df_sales[df_sales['品类'] == category]
        return_cat = df_return[df_return['品类'] == category]
        
        # 老店
        old_sales = sales_cat[sales_cat['是否新店'] == False]
        old_return = return_cat[return_cat['是否新店'] == False]
        old_wholesale = old_sales['销售金额'].sum() - old_return['退回金额'].sum()
        old_profit = (old_sales['销售金额'].sum() - old_sales['暂估成本'].sum()) / 1.13 - \
                    (old_return['退回金额'].sum() - old_return['暂估成本'].sum()) / 1.13
        
        # 新店
        new_sales = sales_cat[sales_cat['是否新店'] == True]
        new_return = return_cat[return_cat['是否新店'] == True]
        new_wholesale = new_sales['销售金额'].sum() - new_return['退回金额'].sum()
        new_profit = (new_sales['销售金额'].sum() - new_sales['暂估成本'].sum()) / 1.13 - \
                    (new_return['退回金额'].sum() - new_return['暂估成本'].sum()) / 1.13
        
        results[category] = {
            'old_wholesale': old_wholesale,
            'new_wholesale': new_wholesale,
            'old_profit': old_profit,
            'new_profit': new_profit
        }
    
    # 黄金外采-新模式
    if df_warehouse is not None and df_other is not None:
        old_warehouse_cost = df_warehouse[df_warehouse['是否新店'] == False]['总成本'].sum()
        old_settlement = df_other[df_other['是否新店'] == False]['发生金额'].sum() / 1.06
        new_warehouse_cost = df_warehouse[df_warehouse['是否新店'] == True]['总成本'].sum()
        new_settlement = df_other[df_other['是否新店'] == True]['发生金额'].sum() / 1.06
        
        results['黄金外采-新模式'] = {
            'old_wholesale': old_warehouse_cost + old_settlement,
            'new_wholesale': new_warehouse_cost + new_settlement,
            'old_profit': old_settlement,
            'new_profit': new_settlement
        }
    else:
        results['黄金外采-新模式'] = {
            'old_wholesale': 0, 'new_wholesale': 0,
            'old_profit': 0, 'new_profit': 0
        }
    
    # 计算合计和总计
    total = {k: sum(results[cat][k] for cat in categories) for k in ['old_wholesale', 'new_wholesale', 'old_profit', 'new_profit']}
    results['合计'] = total
    
    grand_total = {k: total[k] + results['黄金外采-新模式'][k] for k in ['old_wholesale', 'new_wholesale', 'old_profit', 'new_profit']}
    results['总计'] = grand_total
    
    return results


def create_comparison_report(results_2024, results_2025, output_file):
    """生成对比报表"""
    
    print(f"\n📊 生成对比报表...")
    
    categories = ['黄金', '钻石', '爱尚炫', '18K', '翡翠', '其他', '合计', '黄金外采-新模式', '总计']
    
    data = []
    for cat in categories:
        r24 = results_2024[cat]
        r25 = results_2025[cat]
        
        # 老店批发
        old_wholesale_growth = ((r25['old_wholesale'] - r24['old_wholesale']) / r24['old_wholesale'] * 100) if r24['old_wholesale'] != 0 else 0
        old_wholesale_diff = r25['old_wholesale'] - r24['old_wholesale']
        
        # 2025年批发合计
        total_wholesale_2025 = r25['old_wholesale'] + r25['new_wholesale']
        total_wholesale_growth = ((total_wholesale_2025 - r24['old_wholesale']) / r24['old_wholesale'] * 100) if r24['old_wholesale'] != 0 else 0
        
        # 老店毛利
        old_profit_growth = ((r25['old_profit'] - r24['old_profit']) / r24['old_profit'] * 100) if r24['old_profit'] != 0 else 0
        old_profit_diff = r25['old_profit'] - r24['old_profit']
        
        # 2025年毛利合计
        total_profit_2025 = r25['old_profit'] + r25['new_profit']
        total_profit_growth = ((total_profit_2025 - r24['old_profit']) / r24['old_profit'] * 100) if r24['old_profit'] != 0 else 0
        
        data.append({
            '品类': cat,
            '老店批发_2024': r24['old_wholesale'],
            '老店批发_2025': r25['old_wholesale'],
            '老店批发同比': old_wholesale_growth,
            '老店批发差异': old_wholesale_diff,
            '新店批发_2025': r25['new_wholesale'],
            '2025批发合计': total_wholesale_2025,
            '2025批发合计同比': total_wholesale_growth,
            '老店毛利_2024': r24['old_profit'],
            '老店毛利_2025': r25['old_profit'],
            '老店毛利同比': old_profit_growth,
            '老店毛利差异': old_profit_diff,
            '新店毛利_2025': r25['new_profit'],
            '2025毛利合计': total_profit_2025,
            '2025毛利合计同比': total_profit_growth
        })
    
    df = pd.DataFrame(data)
    df.to_excel(output_file, index=False, sheet_name='新老店对比')
    
    # 美化Excel
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    
    # 样式定义
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    subheader_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
    subheader_font = Font(name='微软雅黑', size=10, bold=True)
    normal_font = Font(name='微软雅黑', size=10)
    total_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    total_font = Font(name='微软雅黑', size=10, bold=True)
    border = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
                   top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
    
    # 插入主表头
    ws.insert_rows(1)
    ws.merge_cells('A1:A2')
    ws['A1'] = '品类'
    
    ws.merge_cells('B1:E1')
    ws['B1'] = '老店批发（含税）'
    
    ws.merge_cells('F1:F2')
    ws['F1'] = '新店批发（含税）'
    
    ws.merge_cells('G1:H1')
    ws['G1'] = '2025年新老店合计'
    
    ws.merge_cells('I1:L1')
    ws['I1'] = '老店毛利（未税）'
    
    ws.merge_cells('M1:M2')
    ws['M1'] = '新店毛利（未税）'
    
    ws.merge_cells('N1:O1')
    ws['N1'] = '2025年新老店合计'
    
    # 设置主表头样式
    for cell in ['A1', 'B1', 'F1', 'G1', 'I1', 'M1', 'N1']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        ws[cell].border = border
    
    # 设置子表头样式
    for cell in ws[2]:
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 设置列宽
    widths = {'A': 18, 'B': 15, 'C': 15, 'D': 12, 'E': 15, 'F': 15, 'G': 15, 'H': 12,
              'I': 15, 'J': 15, 'K': 12, 'L': 15, 'M': 15, 'N': 15, 'O': 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    # 设置数据行样式
    for row_idx in range(3, ws.max_row + 1):
        category = ws[f'A{row_idx}'].value
        is_total = category in ['合计', '总计']
        
        for col_idx in range(1, 16):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = total_font if is_total else normal_font
            cell.border = border
            cell.alignment = Alignment(horizontal='right' if col_idx > 1 else 'left', vertical='center')
            
            if is_total:
                cell.fill = total_fill
            
            # 格式化
            if col_idx in [2, 3, 5, 6, 7, 9, 10, 12, 13, 14]:  # 金额列
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
            elif col_idx in [4, 8, 11, 15]:  # 百分比列
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0"%"'
                    if cell.value > 0:
                        cell.font = Font(name='微软雅黑', size=10, color='008000', bold=is_total)
                    elif cell.value < 0:
                        cell.font = Font(name='微软雅黑', size=10, color='FF0000', bold=is_total)
    
    ws.freeze_panes = 'B3'
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    
    # 添加注释
    note_row = ws.max_row + 2
    ws[f'A{note_row}'] = '注：剔除关闭店；移店、更换加盟商视同老店；净批发额、净毛利额数据'
    ws[f'A{note_row}'].font = Font(name='微软雅黑', size=9, italic=True, color='666666')
    
    note_row += 1
    ws[f'A{note_row}'] = '数据说明：2024年全年 vs 2025年1月至今；黄金外采数据已按年份过滤'
    ws[f'A{note_row}'].font = Font(name='微软雅黑', size=9, italic=True, color='666666')
    
    wb.save(output_file)
    print(f"✅ 报表已保存: {output_file}")


if __name__ == "__main__":
    if len(sys.argv) < 8:
        print("用法: python compare_years.py <2024销售> <2024退货> <2025销售> <2025退货> <组织匹配表> <入库单> <其他结算单> [输出文件]")
        print("\n示例:")
        print("  python compare_years.py 销售2024.xlsx 退货2024.xlsx 销售2025.xlsx 退货2025.xlsx 组织匹配.xlsx 入库单.xlsx 其他结算单.xlsx")
        sys.exit(1)
    
    sales_2024 = sys.argv[1]
    return_2024 = sys.argv[2]
    sales_2025 = sys.argv[3]
    return_2025 = sys.argv[4]
    org_file = sys.argv[5]
    warehouse_file = sys.argv[6]
    other_file = sys.argv[7]
    output_file = sys.argv[8] if len(sys.argv) > 8 else "2024-2025年加盟新老店对比分析.xlsx"
    
    # 读取组织匹配表
    print("📊 读取组织匹配表...")
    org_mapping_df = pd.read_excel(org_file)
    org_mapping_df['组织'] = org_mapping_df['加盟门店'].str.replace(r'[*]+$', '', regex=True)
    
    org_mapping = {
        'new_stores': set(org_mapping_df[org_mapping_df['2025年新店'] == '是']['组织'].values),
        'closed_stores': set(org_mapping_df[org_mapping_df['已闭店'] == '是']['组织'].values)
    }
    
    print(f"   2025年新店: {len(org_mapping['new_stores'])}个")
    print(f"   已闭店: {len(org_mapping['closed_stores'])}个")
    
    # 分析2024年
    results_2024 = analyze_year(2024, sales_2024, return_2024, org_mapping, 
                                warehouse_file, other_file)
    
    # 分析2025年
    results_2025 = analyze_year(2025, sales_2025, return_2025, org_mapping,
                                warehouse_file, other_file)
    
    # 生成对比报表
    create_comparison_report(results_2024, results_2025, output_file)
    
    print(f"\n✅ 对比分析完成！")
