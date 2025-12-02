#!/usr/bin/env python3
"""
美化Excel输出格式
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def format_beautiful_excel(input_file, output_file, start_date="2025-01-01", end_date="2025-11-30"):
    """
    将分析结果格式化为美观的Excel文件

    参数:
        input_file: 输入的Excel文件
        output_file: 输出的Excel文件
        start_date: 开始日期
        end_date: 结束日期
    """
    # 读取数据
    df = pd.read_excel(input_file)

    # 创建新工作簿
    wb = Workbook()
    ws = wb.active

    # 定义样式
    title_font = Font(name='微软雅黑', size=14, bold=True)
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='right', vertical='center')
    category_alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    red_font = Font(name='微软雅黑', size=10, color='FF0000')

    # 1. 写标题
    title_text = f"{start_date.replace('-', '年', 1).replace('-', '月')}日～{end_date.replace('-', '年', 1).replace('-', '月')}日加盟新老店批发&毛利情况"
    ws['A1'] = title_text
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:O1')
    ws.row_dimensions[1].height = 25

    # 2. 写表头第2行（主分类）
    ws['A2'] = '品类'
    ws['B2'] = '老店批发（含税）'
    ws['F2'] = '新店批发（含税）'
    ws['G2'] = '2025年新老店合计'
    ws['I2'] = '老店毛利（未税）'
    ws['M2'] = '新店毛利（未税）'
    ws['N2'] = '2025年新老店合计'

    # 3. 写表头第3行（子分类）
    headers_row3 = ['品类', '2024年', '2025年', '同比增长率', '同比差异', '2025年', '批发额', '同比增长率',
                    '2024年', '2025年', '同比增长率', '同比差异', '2025年', '毛利额', '同比增长率']

    for col_idx, header in enumerate(headers_row3, start=1):
        ws.cell(row=3, column=col_idx, value=header)

    # 4. 应用表头样式
    for col_idx in range(1, 16):
        # 第2行
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

        # 第3行
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 5. 合并表头单元格
    ws.merge_cells('A2:A3')  # 品类
    ws.merge_cells('B2:E2')  # 老店批发（含税）
    ws.merge_cells('F2:F3')  # 新店批发（含税）
    ws.merge_cells('G2:H2')  # 2025年新老店合计
    ws.merge_cells('I2:L2')  # 老店毛利（未税）
    ws.merge_cells('M2:M3')  # 新店毛利（未税）
    ws.merge_cells('N2:O2')  # 2025年新老店合计

    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 25

    # 6. 写入数据
    for row_idx, row_data in df.iterrows():
        excel_row = row_idx + 4  # 从第4行开始

        # 品类
        ws.cell(row=excel_row, column=1, value=row_data['品类'])

        # 老店批发（含税）
        ws.cell(row=excel_row, column=2, value=row_data['老店批发（含税）_2024年'])
        ws.cell(row=excel_row, column=3, value=row_data['老店批发（含税）_2025年'])
        ws.cell(row=excel_row, column=4, value=row_data['老店批发（含税）_同比增长率'])
        ws.cell(row=excel_row, column=5, value=row_data['老店批发（含税）_同比差异'])

        # 新店批发（含税）
        ws.cell(row=excel_row, column=6, value=row_data['新店批发（含税）_2025年'])

        # 2025年新老店合计
        ws.cell(row=excel_row, column=7, value=row_data['2025年新老店合计_批发额'])
        ws.cell(row=excel_row, column=8, value=row_data['2025年新老店合计_同比增长率'])

        # 老店毛利（未税）
        ws.cell(row=excel_row, column=9, value=row_data['老店毛利（未税）_2024年'])
        ws.cell(row=excel_row, column=10, value=row_data['老店毛利（未税）_2025年'])
        ws.cell(row=excel_row, column=11, value=row_data['老店毛利（未税）_同比增长率'])
        ws.cell(row=excel_row, column=12, value=row_data['老店毛利（未税）_同比差异'])

        # 新店毛利（未税）
        ws.cell(row=excel_row, column=13, value=row_data['新店毛利（未税）_2025年'])

        # 2025年新老店合计
        ws.cell(row=excel_row, column=14, value=row_data['2025年新老店合计_毛利额'])
        ws.cell(row=excel_row, column=15, value=row_data['2025年新老店合计_同比增长率_毛利'])

    # 7. 格式化数据行
    for row_idx in range(4, 4 + len(df)):
        for col_idx in range(1, 16):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

            if col_idx == 1:  # 品类列
                cell.alignment = category_alignment
                cell.font = data_font
                # 合计和总计行加粗
                if cell.value in ['合计', '总计', '黄金外采-新模式']:
                    cell.font = Font(name='微软雅黑', size=10, bold=True)
            else:
                cell.alignment = data_alignment
                cell.font = data_font

                # 格式化数字
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    # 同比增长率列
                    if col_idx in [4, 8, 11, 15]:
                        cell.number_format = '0.0%'
                        if cell.value < 0:
                            cell.font = red_font
                    else:
                        cell.number_format = '#,##0'
                        # 同比差异列
                        if col_idx in [5, 12] and cell.value < 0:
                            cell.font = red_font

    # 8. 设置列宽
    column_widths = {
        'A': 15, 'B': 12, 'C': 12, 'D': 12, 'E': 12,
        'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12,
        'K': 12, 'L': 12, 'M': 12, 'N': 12, 'O': 12,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 9. 添加底部注释
    note_row = 4 + len(df) + 2
    ws[f'A{note_row}'] = '注：剔除关闭店;移店、更换加盟商视同老店;净批发额、净毛利额数据。'
    ws[f'A{note_row}'].font = Font(name='微软雅黑', size=9, italic=True)
    ws[f'A{note_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{note_row}:O{note_row}')

    # 保存
    wb.save(output_file)
    print(f"✅ 已生成美化的Excel文件: {output_file}")


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("用法: python format_beautiful_excel.py <输入文件> <输出文件> [开始日期] [结束日期]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]
    start_date = sys.argv[3] if len(sys.argv) > 3 else "2025-01-01"
    end_date = sys.argv[4] if len(sys.argv) > 4 else "2025-11-30"

    format_beautiful_excel(input_file, output_file, start_date, end_date)
