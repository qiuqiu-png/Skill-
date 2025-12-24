#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
货盘需求计算器
根据库存、在途、销售数据自动计算门店补货需求量
支持普通商品和DIY商品的差异化计算规则
"""

import sys
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


def calculate_goods_demand(input_file, output_file=None):
    """
    计算货盘需求

    参数:
        input_file: 输入Excel文件路径
        output_file: 输出Excel文件路径（可选，默认添加_处理后后缀）
    """
    # 如果没有指定输出文件，自动生成
    if output_file is None:
        if input_file.endswith('.xlsx'):
            output_file = input_file.replace('.xlsx', '_处理后.xlsx')
        else:
            output_file = input_file + '_处理后.xlsx'

    print("正在加载Excel文件...")
    wb = load_workbook(input_file, keep_vba=False, data_only=False, keep_links=False)
    ws = wb.active

    print(f"工作表名称: {ws.title}")
    print(f"总行数: {ws.max_row}, 总列数: {ws.max_column}")

    # 定义列索引
    b_col = column_index_from_string('B')   # 大类
    e_col = column_index_from_string('E')   # 二级分类
    g_col = column_index_from_string('G')   # 周期类型
    ci_col = column_index_from_string('CI')  # 补单数量
    ds_col = column_index_from_string('DS')  # 门店数据开始列
    yd_col = column_index_from_string('YD')  # 门店数据结束列

    # 第一步：获取第6行的门店分组（每个门店占4列：库存、销售、在途、需求）
    print("\n=== 第一步：识别门店分组 ===")
    stores = []
    current_col = ds_col

    while current_col <= yd_col:
        store_name = ws.cell(6, current_col).value
        if store_name:
            # 检查第7行的列标题是否符合预期
            h1 = ws.cell(7, current_col).value
            h2 = ws.cell(7, current_col + 1).value
            h3 = ws.cell(7, current_col + 2).value
            h4 = ws.cell(7, current_col + 3).value

            if h1 == '库存' and h2 == '销售' and h3 == '在途' and h4 == '需求':
                stores.append({
                    'name': store_name,
                    'inventory_col': current_col,      # 库存
                    'sales_col': current_col + 1,      # 销售
                    'intransit_col': current_col + 2,  # 在途
                    'demand_col': current_col + 3      # 需求
                })
                current_col += 4
            else:
                current_col += 1
        else:
            current_col += 1

    print(f"找到 {len(stores)} 个门店")
    if stores:
        print(f"门店示例: {stores[0]['name']}, {stores[1]['name'] if len(stores) > 1 else ''}")

    # 第二步：筛选符合条件的行并计算需求
    print("\n=== 第二步：处理数据 ===")
    print("筛选条件: B列=足金/足金（新）、G列=21天周期、CI列补单数>0")

    processed_count = 0
    skipped_rows = 0

    for row_idx in range(8, ws.max_row + 1):
        # 获取筛选条件列的值
        b_val = ws.cell(row_idx, b_col).value
        g_val = ws.cell(row_idx, g_col).value
        ci_val = ws.cell(row_idx, ci_col).value

        # 筛选条件1：B列=足金或足金（新）
        if b_val not in ['足金', '足金（新）']:
            continue

        # 筛选条件2：G列=21天周期
        if g_val != '21天周期':
            continue

        # 筛选条件3：CI列补单数>0
        try:
            ci_num = float(ci_val) if ci_val is not None else 0
        except:
            ci_num = 0

        if ci_num <= 0:
            continue

        # 获取E列二级分类（判断是否为DIY）
        e_val = ws.cell(row_idx, e_col).value
        is_diy = (e_val == 'DIY')

        # 对每个门店处理
        for store in stores:
            # 获取库存、销售、在途
            inventory = ws.cell(row_idx, store['inventory_col']).value
            sales = ws.cell(row_idx, store['sales_col']).value
            intransit = ws.cell(row_idx, store['intransit_col']).value

            # 转换为数值
            try:
                inv = float(inventory) if inventory is not None else 0
            except:
                inv = 0

            try:
                sal = float(sales) if sales is not None else 0
            except:
                sal = 0

            try:
                itr = float(intransit) if intransit is not None else 0
            except:
                itr = 0

            # 忽略销售<=0或为空的门店
            if sal <= 0:
                continue

            # 计算库存+在途-销售
            calc = inv + itr - sal

            # 根据规则填充需求列
            demand_value = None

            if is_diy:
                # DIY商品的特殊规则
                if calc < 0:
                    # 库存+在途-销售 < 0
                    if ci_num > 500:
                        demand_value = abs(calc) * 5
                    else:
                        demand_value = abs(calc) * 2
                elif calc == 0:
                    # 库存+在途-销售 = 0
                    demand_value = 1
                else:
                    # 库存+在途-销售 > 0，且销售>1的，填1
                    if sal > 1:
                        demand_value = 1
            else:
                # 普通商品的规则
                if calc < 0:
                    demand_value = abs(calc)
                elif calc == 0:
                    demand_value = 1
                # calc > 0 时不填充

            # 填充需求列
            if demand_value is not None:
                ws.cell(row_idx, store['demand_col']).value = demand_value
                processed_count += 1

        # 进度提示
        if row_idx % 500 == 0:
            print(f"  已处理到第 {row_idx} 行...")

    print(f"\n✅ 处理完成！")
    print(f"📊 共填充了 {processed_count} 个需求单元格")

    # 保存文件
    print(f"\n正在保存文件到: {output_file}")
    wb.save(output_file)
    wb.close()

    print(f"✅ 文件已保存！")

    return output_file


def main():
    """命令行入口"""
    if len(sys.argv) < 2:
        print("用法: python3 calculate_demand.py <输入文件路径> [输出文件路径]")
        print("示例: python3 calculate_demand.py /path/to/货盘数据.xlsx")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        calculate_goods_demand(input_file, output_file)
    except Exception as e:
        print(f"\n❌ 处理失败: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
