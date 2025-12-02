#!/usr/bin/env python3
"""
分析2025年新店的批发和毛利明细
按每个新店单独统计,包含开业时间和按件/按克批发额
"""

import sys
import pandas as pd
from pathlib import Path


def analyze_new_stores_detail(sales_file, return_file, org_mapping_file, output_path=None):
    """
    分析2025年新店的批发和毛利明细

    参数:
        sales_file: 销售表格文件路径
        return_file: 退货表格文件路径
        org_mapping_file: 组织匹配表文件路径
        output_path: 输出文件路径(可选)
    """
    try:
        print("📊 读取组织匹配表...")
        org_mapping = pd.read_excel(org_mapping_file)
        print(f"   组织数量: {len(org_mapping)}")

        # 处理组织名称,去除末尾的*和M
        org_mapping['组织'] = org_mapping['加盟门店'].str.replace(r'[*M]+$', '', regex=True)

        # 创建新店和已闭店字典
        new_stores = set(org_mapping[org_mapping['2025年新店'] == '是']['组织'].values)
        closed_stores = set(org_mapping[org_mapping['已闭店'] == '是']['组织'].values)

        # 创建开业日期映射字典
        opening_date_map = {}
        if '开业日期' in org_mapping.columns:
            for _, row in org_mapping.iterrows():
                opening_date_map[row['组织']] = row['开业日期']

        print(f"   2025年新店数量: {len(new_stores)}")
        print(f"   已闭店数量: {len(closed_stores)}")

        # 读取销售数据
        print("\n📊 读取销售数据...")
        df_sales = pd.read_excel(sales_file, sheet_name=0)
        print(f"   原始数据行数: {len(df_sales)}")

        # 只保留组织名称以M结尾的门店
        df_sales = df_sales[df_sales['组织'].str.rstrip().str.endswith('M', na=False)]
        print(f"   仅保留以M结尾的门店后行数: {len(df_sales)}")

        # 处理组织名称
        df_sales['组织_cleaned'] = df_sales['组织'].str.replace(r'[*M]+$', '', regex=True)

        # 剔除已闭店
        df_sales = df_sales[~df_sales['组织_cleaned'].isin(closed_stores)]
        print(f"   剔除闭店后行数: {len(df_sales)}")

        # 只保留新店
        df_sales = df_sales[df_sales['组织_cleaned'].isin(new_stores)]
        print(f"   仅保留新店后行数: {len(df_sales)}")

        # 读取退货数据
        print("\n📊 读取退货数据...")
        df_return = pd.read_excel(return_file, sheet_name=0)
        print(f"   原始数据行数: {len(df_return)}")

        # 只保留组织名称以M结尾的门店
        df_return = df_return[df_return['组织'].str.rstrip().str.endswith('M', na=False)]
        print(f"   仅保留以M结尾的门店后行数: {len(df_return)}")

        # 处理组织名称
        df_return['组织_cleaned'] = df_return['组织'].str.replace(r'[*M]+$', '', regex=True)

        # 剔除已闭店
        df_return = df_return[~df_return['组织_cleaned'].isin(closed_stores)]
        print(f"   剔除闭店后行数: {len(df_return)}")

        # 只保留新店
        df_return = df_return[df_return['组织_cleaned'].isin(new_stores)]
        print(f"   仅保留新店后行数: {len(df_return)}")

        # 确保数值列为数字类型
        df_sales['销售金额'] = pd.to_numeric(df_sales['销售金额'], errors='coerce').fillna(0)
        df_sales['暂估成本'] = pd.to_numeric(df_sales['暂估成本'], errors='coerce').fillna(0)
        df_return['退回金额'] = pd.to_numeric(df_return['退回金额'], errors='coerce').fillna(0)
        df_return['暂估成本'] = pd.to_numeric(df_return['暂估成本'], errors='coerce').fillna(0)

        print("\n📊 开始按新店统计...")

        # 获取所有新店列表(从组织匹配表,包括没有销售数据的新店)
        new_store_list = sorted(list(new_stores))

        results = []

        for store in new_store_list:
            # 该店的销售数据
            store_sales = df_sales[df_sales['组织_cleaned'] == store]
            store_return = df_return[df_return['组织_cleaned'] == store]

            # 获取开业日期
            opening_date = opening_date_map.get(store, None)
            if pd.notna(opening_date):
                if isinstance(opening_date, pd.Timestamp):
                    opening_date_str = opening_date.strftime('%Y-%m-%d')
                else:
                    opening_date_str = str(opening_date)
            else:
                opening_date_str = ''

            # 销售金额
            sales_amount = store_sales['销售金额'].sum()

            # 退回金额
            return_amount = store_return['退回金额'].sum()

            # 批发额(含税)= 销售金额 - 退回金额
            wholesale = sales_amount - return_amount

            # 按件批发额:计量单位='件'
            store_sales_piece = store_sales[store_sales['计量单位'] == '件']
            store_return_piece = store_return[store_return['计量单位'] == '件']
            wholesale_piece = store_sales_piece['销售金额'].sum() - store_return_piece['退回金额'].sum()

            # 按克批发额:计量单位='克'
            store_sales_gram = store_sales[store_sales['计量单位'] == '克']
            store_return_gram = store_return[store_return['计量单位'] == '克']
            wholesale_gram = store_sales_gram['销售金额'].sum() - store_return_gram['退回金额'].sum()

            # 毛利额(未税)= (销售金额 - 暂估成本)/1.13 - (退回金额 - 暂估成本)/1.13
            profit = (store_sales['销售金额'].sum() - store_sales['暂估成本'].sum()) / 1.13 - \
                    (store_return['退回金额'].sum() - store_return['暂估成本'].sum()) / 1.13

            # 毛利率 = 毛利额 / 批发额
            profit_rate = (profit / wholesale * 100) if wholesale != 0 else 0

            results.append({
                '新店组织名称': store,
                '开业时间': opening_date_str,
                '销售金额': sales_amount,
                '退货金额': return_amount,
                '批发额(含税)': wholesale,
                '按件批发额(含税)': wholesale_piece,
                '按克批发额(含税)': wholesale_gram,
                '毛利额(未税)': profit,
                '毛利率(%)': profit_rate
            })

        # 创建结果DataFrame
        result_df = pd.DataFrame(results)

        # 添加合计行
        total_row = {
            '新店组织名称': '合计',
            '开业时间': '',
            '销售金额': result_df['销售金额'].sum(),
            '退货金额': result_df['退货金额'].sum(),
            '批发额(含税)': result_df['批发额(含税)'].sum(),
            '按件批发额(含税)': result_df['按件批发额(含税)'].sum(),
            '按克批发额(含税)': result_df['按克批发额(含税)'].sum(),
            '毛利额(未税)': result_df['毛利额(未税)'].sum(),
            '毛利率(%)': (result_df['毛利额(未税)'].sum() / result_df['批发额(含税)'].sum() * 100)
                        if result_df['批发额(含税)'].sum() != 0 else 0
        }
        result_df = pd.concat([result_df, pd.DataFrame([total_row])], ignore_index=True)

        # 打印结果
        print("\n" + "=" * 180)
        print("📈 2025年新店批发&毛利明细")
        print("=" * 180)
        print()

        # 格式化输出
        for _, row in result_df.iterrows():
            print(f"{row['新店组织名称']:40s}  "
                  f"开业: {row['开业时间']:12s}  "
                  f"销售: {row['销售金额']:>13,.0f}  "
                  f"退货: {row['退货金额']:>13,.0f}  "
                  f"批发: {row['批发额(含税)']:>13,.0f}  "
                  f"按件: {row['按件批发额(含税)']:>13,.0f}  "
                  f"按克: {row['按克批发额(含税)']:>13,.0f}  "
                  f"毛利: {row['毛利额(未税)']:>13,.0f}  "
                  f"毛利率: {row['毛利率(%)']:>6.2f}%")

        print()

        # 保存到Excel
        if output_path:
            output_file = output_path
        else:
            output_file = "2025年新店批发毛利明细.xlsx"

        # 使用openpyxl美化输出
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "新店明细"

        # 定义样式
        title_font = Font(name='微软雅黑', size=14, bold=True)
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        data_font = Font(name='微软雅黑', size=10)
        data_alignment_right = Alignment(horizontal='right', vertical='center')
        data_alignment_left = Alignment(horizontal='left', vertical='center')
        data_alignment_center = Alignment(horizontal='center', vertical='center')

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # 标题
        ws['A1'] = '2025年新店批发&毛利明细'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:I1')
        ws.row_dimensions[1].height = 25

        # 表头
        headers = ['新店组织名称', '开业时间', '销售金额', '退货金额', '批发额(含税)',
                   '按件批发额(含税)', '按克批发额(含税)', '毛利额(未税)', '毛利率(%)']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[2].height = 25

        # 写入数据
        for row_idx, row_data in result_df.iterrows():
            excel_row = row_idx + 3

            # 新店组织名称
            ws.cell(row=excel_row, column=1, value=row_data['新店组织名称'])
            ws.cell(row=excel_row, column=1).alignment = data_alignment_left
            ws.cell(row=excel_row, column=1).border = thin_border
            ws.cell(row=excel_row, column=1).font = data_font

            # 合计行加粗
            if row_data['新店组织名称'] == '合计':
                ws.cell(row=excel_row, column=1).font = Font(name='微软雅黑', size=10, bold=True)

            # 开业时间
            ws.cell(row=excel_row, column=2, value=row_data['开业时间'])
            ws.cell(row=excel_row, column=2).alignment = data_alignment_center
            ws.cell(row=excel_row, column=2).border = thin_border
            ws.cell(row=excel_row, column=2).font = data_font

            # 销售金额
            ws.cell(row=excel_row, column=3, value=row_data['销售金额'])
            ws.cell(row=excel_row, column=3).number_format = '#,##0'
            ws.cell(row=excel_row, column=3).alignment = data_alignment_right
            ws.cell(row=excel_row, column=3).border = thin_border
            ws.cell(row=excel_row, column=3).font = data_font

            # 退货金额
            ws.cell(row=excel_row, column=4, value=row_data['退货金额'])
            ws.cell(row=excel_row, column=4).number_format = '#,##0'
            ws.cell(row=excel_row, column=4).alignment = data_alignment_right
            ws.cell(row=excel_row, column=4).border = thin_border
            ws.cell(row=excel_row, column=4).font = data_font

            # 批发额
            ws.cell(row=excel_row, column=5, value=row_data['批发额(含税)'])
            ws.cell(row=excel_row, column=5).number_format = '#,##0'
            ws.cell(row=excel_row, column=5).alignment = data_alignment_right
            ws.cell(row=excel_row, column=5).border = thin_border
            ws.cell(row=excel_row, column=5).font = data_font

            # 按件批发额
            ws.cell(row=excel_row, column=6, value=row_data['按件批发额(含税)'])
            ws.cell(row=excel_row, column=6).number_format = '#,##0'
            ws.cell(row=excel_row, column=6).alignment = data_alignment_right
            ws.cell(row=excel_row, column=6).border = thin_border
            ws.cell(row=excel_row, column=6).font = data_font

            # 按克批发额
            ws.cell(row=excel_row, column=7, value=row_data['按克批发额(含税)'])
            ws.cell(row=excel_row, column=7).number_format = '#,##0'
            ws.cell(row=excel_row, column=7).alignment = data_alignment_right
            ws.cell(row=excel_row, column=7).border = thin_border
            ws.cell(row=excel_row, column=7).font = data_font

            # 毛利额
            ws.cell(row=excel_row, column=8, value=row_data['毛利额(未税)'])
            ws.cell(row=excel_row, column=8).number_format = '#,##0'
            ws.cell(row=excel_row, column=8).alignment = data_alignment_right
            ws.cell(row=excel_row, column=8).border = thin_border
            ws.cell(row=excel_row, column=8).font = data_font

            # 毛利率
            ws.cell(row=excel_row, column=9, value=row_data['毛利率(%)'] / 100)
            ws.cell(row=excel_row, column=9).number_format = '0.00%'
            ws.cell(row=excel_row, column=9).alignment = data_alignment_right
            ws.cell(row=excel_row, column=9).border = thin_border
            ws.cell(row=excel_row, column=9).font = data_font

        # 设置列宽
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 12

        # 添加底部注释
        note_row = len(result_df) + 4
        ws[f'A{note_row}'] = '注:剔除关闭店;仅统计以M结尾的门店;净批发额、净毛利额数据;按件/按克根据计量单位区分。'
        ws[f'A{note_row}'].font = Font(name='微软雅黑', size=9, italic=True)
        ws[f'A{note_row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'A{note_row}:I{note_row}')

        # 保存
        wb.save(output_file)
        print(f"✅ 结果已保存到: {output_file}")
        print(f"   注:剔除关闭店;仅统计以M结尾的门店;净批发额、净毛利额数据")

        return result_df

    except Exception as e:
        print(f"❌ 错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("用法: python analyze_new_stores_detail.py <销售文件> <退货文件> <组织匹配表> [输出文件]")
        print("\n示例:")
        print("  python analyze_new_stores_detail.py 销售.xlsx 退货.xlsx 组织匹配.xlsx")
        print("  python analyze_new_stores_detail.py 销售.xlsx 退货.xlsx 组织匹配.xlsx 新店明细.xlsx")
        sys.exit(1)

    sales_file = sys.argv[1]
    return_file = sys.argv[2]
    org_mapping = sys.argv[3]
    output_file = sys.argv[4] if len(sys.argv) > 4 else None

    analyze_new_stores_detail(sales_file, return_file, org_mapping, output_file)
