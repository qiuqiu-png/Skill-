#!/usr/bin/env python3
"""
款式销售排名分析工具
根据销售明细数据生成款式销售排名报表
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os


class StyleSalesRanking:
    """款式销售排名分析类"""

    def __init__(self, sales_file_path):
        """
        初始化

        Args:
            sales_file_path: 销售明细Excel文件路径
        """
        self.sales_file_path = sales_file_path
        self.df = None
        self.style_sales = None
        self.categories = []
        self.category_rules = []

        # 样式定义
        self.header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.header_font = Font(bold=True, size=11)
        self.title_font = Font(bold=True, size=14)
        self.link_font = Font(color="0563C1", underline="single")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def load_data(self):
        """加载销售数据"""
        print(f'正在读取数据: {self.sales_file_path}')
        self.df = pd.read_excel(self.sales_file_path)

        # 只保留销售记录
        self.df = self.df[self.df['销售'] == '销售'].copy()
        print(f'销售记录数: {len(self.df)}')

        # 排除"笼统款"
        before_count = len(self.df)
        self.df = self.df[~self.df['款式'].str.contains('笼统款', na=False)].copy()
        after_count = len(self.df)
        print(f'已排除"笼统款": {before_count - after_count}条, 剩余{after_count}条')

        return self.df

    def load_category_rules(self, rules_file_path):
        """
        加载分类判定条件

        Args:
            rules_file_path: 分类判定条件Excel文件路径
        """
        print(f'\n读取分类判定条件: {rules_file_path}')
        rules_df = pd.read_excel(rules_file_path, header=None)

        # 提取分类规则
        self.category_rules = []
        for idx, row in rules_df.iterrows():
            if pd.notna(row[1]) and str(row[1]).isdigit():
                seq = int(row[1])
                category = row[2]
                condition = row[3]
                self.category_rules.append({
                    'seq': seq,
                    'category': category,
                    'condition': condition
                })

        print(f'已加载 {len(self.category_rules)} 个分类规则')
        for rule in self.category_rules:
            print(f"  {rule['seq']}. {rule['category']}: {rule['condition']}")

        return self.category_rules

    def classify_item(self, row):
        """
        对单条记录进行分类

        Args:
            row: DataFrame的一行数据

        Returns:
            分类名称
        """
        # 1. 黄金戒指系列
        if pd.notna(row['三级分类']) and '戒指' in str(row['三级分类']):
            return '黄金戒指系列'

        # 2. 凤华&凤华2.0
        if pd.notna(row['主题名称']) and '凤华' in str(row['主题名称']):
            return '凤华&凤华2.0'

        # 3. 恋恋风情
        if pd.notna(row['三级分类']) and str(row['三级分类']) == '恋恋风情':
            return '恋恋风情'

        # 4. 点钻
        if pd.notna(row['主题名称']) and '繁花' in str(row['主题名称']):
            return '点钻'

        # 5. 新奢
        if pd.notna(row['款式']) and str(row['款式']).startswith('S-'):
            return '新奢'

        # 6. 暮光之城
        if pd.notna(row['主题名称']) and '暮光之城' in str(row['主题名称']):
            return '暮光之城'

        # 7. 古韵传香
        if pd.notna(row['主题名称']) and '古韵传香' in str(row['主题名称']):
            return '古韵传香'

        # 8. 锦绣金
        if pd.notna(row['三级分类']) and str(row['三级分类']) == '锦绣金':
            return '锦绣金'

        # 17. 爱情灵药
        if pd.notna(row['主题名称']) and '爱情灵药' in str(row['主题名称']):
            return '爱情灵药'

        # 18. 娇玉系列
        if pd.notna(row['主题名称']) and '娇玉系列' in str(row['主题名称']):
            return '娇玉系列'

        # 19. 抖音到店
        if pd.notna(row['主题名称']) and '抖音' in str(row['主题名称']):
            return '抖音到店'

        # 20. 盗墓笔记
        if pd.notna(row['主题名称']) and '盗墓' in str(row['主题名称']):
            return '盗墓笔记'

        # 21. 中药
        if pd.notna(row['主题名称']) and ('中药' in str(row['主题名称']) or '本草' in str(row['主题名称'])):
            return '中药'

        # 14. Less系列
        if pd.notna(row['主题名称']) and 'Less' in str(row['主题名称']):
            return 'Less系列'

        # 15. 彩宝
        if pd.notna(row['主题名称']) and '彩宝' in str(row['主题名称']):
            return '彩宝'

        # 11. DIY
        if pd.notna(row['款式大类']) and str(row['款式大类']) == 'DIY转运珠':
            return 'DIY'

        # 判断是否是足金/爱尚金
        is_gold = row['大类'] in ['足金', '足金(新)', '爱尚金']
        is_ke = str(row['计量单位']) == '克'
        is_jian = str(row['计量单位']) == '件'
        has_qy_jm = pd.notna(row['主题名称']) and ('QY' in str(row['主题名称']) or 'JM' in str(row['主题名称']))

        # 10. 其他按克黄金(加盟专款)
        if is_gold and is_ke and has_qy_jm:
            return '其他按克黄金(加盟专款)'

        # 13. 按件黄金(加盟专款)
        if is_jian and has_qy_jm:
            return '按件黄金(加盟专款)'

        # 9. 其他按克黄金(公司款)
        if is_gold and is_ke:
            return '其他按克黄金(公司款)'

        # 12. 按件黄金(公司款)
        if is_jian and not has_qy_jm:
            return '按件黄金(公司款)'

        # 16. 其他K金&铂金镶嵌
        if row['大类'] in ['18K', 'K金镶嵌(新)', '爱尚炫', '钻石', 'Pt950', 'Pt900']:
            return '其他K金&铂金镶嵌'

        return '未分类'

    def apply_classification(self):
        """应用分类规则"""
        print('\n正在应用分类规则...')
        self.df['新分类'] = self.df.apply(self.classify_item, axis=1)

        # 统计分类
        category_stats = self.df['新分类'].value_counts()
        print(f'\n分类统计:')
        for cat, count in category_stats.items():
            print(f'  {cat}: {count}条')

        return category_stats

    def calculate_style_sales(self):
        """统计每个款式的销售件数"""
        print('\n正在统计款式销售...')
        self.style_sales = self.df.groupby(['新分类', '主题名称', '大类', '小类', '款式']).agg({
            '销售件数': 'sum'
        }).reset_index()

        self.style_sales.columns = ['分类', '主题', '大类', '小类', '款式', '销量']

        # 计算分类内排名
        self.style_sales['分类内排名'] = self.style_sales.groupby('分类')['销量'].rank(
            method='dense', ascending=False
        ).astype(int)

        # 排序
        self.style_sales = self.style_sales.sort_values(['分类', '分类内排名'])

        print(f'总计 {len(self.style_sales)} 个不同款式')
        print(f'分为 {self.style_sales["分类"].nunique()} 个分类')

        return self.style_sales

    def generate_cross_store_report(self, output_path):
        """
        生成跨店铺款式排名报表(多sheet)

        Args:
            output_path: 输出文件路径
        """
        print('\n开始生成跨店铺款式排名报表...')

        wb = Workbook()
        wb.remove(wb.active)

        # 创建目录sheet
        catalog_ws = wb.create_sheet('目录', 0)
        catalog_ws.column_dimensions['B'].width = 25
        catalog_ws.column_dimensions['C'].width = 12
        catalog_ws.column_dimensions['D'].width = 15

        catalog_ws['B2'] = '目录'
        catalog_ws['B2'].font = self.title_font

        catalog_ws['B4'] = '序号'
        catalog_ws['C4'] = '分类'
        catalog_ws['D4'] = '链接'
        catalog_ws['E4'] = '备注'

        for col in ['B', 'C', 'D', 'E']:
            cell = catalog_ws[f'{col}4']
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 分类顺序
        category_order = [
            '黄金戒指系列', '凤华&凤华2.0', '恋恋风情', '点钻', '新奢',
            '暮光之城', '古韵传香', '锦绣金', '其他按克黄金(公司款)',
            '其他按克黄金(加盟专款)', 'DIY', '按件黄金(公司款)',
            '按件黄金(加盟专款)', 'Less系列', '彩宝', '其他K金&铂金镶嵌',
            '爱情灵药', '娇玉系列', '抖音到店', '盗墓笔记', '中药'
        ]

        # 只保留有数据的分类
        self.categories = [cat for cat in category_order if cat in self.style_sales['分类'].values]

        print(f'将为 {len(self.categories)} 个分类创建sheet')

        # 创建每个分类的sheet
        for idx, category in enumerate(self.categories, 1):
            print(f'  处理分类 {idx}/{len(self.categories)}: {category}')

            # 添加到目录
            catalog_ws[f'B{idx+4}'] = idx
            catalog_ws[f'C{idx+4}'] = category
            catalog_ws[f'D{idx+4}'] = '进入'

            # 添加备注
            if category in ['黄金戒指系列', '凤华&凤华2.0', '娇玉系列', '恋恋风情', '点钻', 'DIY', 'Less系列']:
                catalog_ws[f'E{idx+4}'] = '★重点关注'
            elif category in ['古韵传香']:
                catalog_ws[f'E{idx+4}'] = '★关注'

            for col in ['B', 'C', 'D', 'E']:
                catalog_ws[f'{col}{idx+4}'].border = self.border
                catalog_ws[f'{col}{idx+4}'].alignment = Alignment(horizontal='center', vertical='center')

            # 获取该分类的数据
            cat_data = self.style_sales[self.style_sales['分类'] == category].copy()

            # 创建sheet
            sheet_name = category[:31] if len(category) > 31 else category
            try:
                ws = wb.create_sheet(sheet_name)
            except:
                sheet_name = f'分类{idx}'
                ws = wb.create_sheet(sheet_name)

            # 设置列宽
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 35
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 12

            # 标题
            ws['B2'] = f'{category}销售排名'
            ws['B2'].font = self.title_font

            ws['G2'] = '返回目录'
            ws['G2'].font = self.link_font
            ws['G2'].alignment = Alignment(horizontal='right')

            # 表头
            headers = ['销售排名', '大类', '小类', '款式', '主题', '销量']
            for col_idx, header in enumerate(headers, 2):
                cell = ws.cell(row=4, column=col_idx)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 填充数据
            for row_idx, (_, row) in enumerate(cat_data.iterrows(), 5):
                ws.cell(row=row_idx, column=2, value=row['分类内排名'])
                ws.cell(row=row_idx, column=3, value=row['大类'])
                ws.cell(row=row_idx, column=4, value=row['小类'])
                ws.cell(row=row_idx, column=5, value=row['款式'])
                ws.cell(row=row_idx, column=6, value=row['主题'])
                ws.cell(row=row_idx, column=7, value=row['销量'])

                for col_idx in range(2, 8):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 保存文件
        wb.save(output_path)
        print(f'\n✅ 报表已保存: {output_path}')
        print(f'生成了 {len(self.categories)} 个分类的销售排名')

        return output_path


def main():
    """主函数 - 示例用法"""
    # 示例: 创建分析对象
    analyzer = StyleSalesRanking('/path/to/销售明细.xlsx')

    # 加载数据
    analyzer.load_data()

    # 加载分类规则(可选)
    # analyzer.load_category_rules('/path/to/分类判定条件.xlsx')

    # 应用分类
    analyzer.apply_classification()

    # 统计销售
    analyzer.calculate_style_sales()

    # 生成报表
    output_path = os.path.expanduser('~/Desktop/门店款式销售排名.xlsx')
    analyzer.generate_cross_store_report(output_path)


if __name__ == '__main__':
    main()
